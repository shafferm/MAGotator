import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def generate_multi_sheet_xlsx(input_file, rrna_file, trna_file, combined_annotations, combined_rrna, output_file):
    # Read the data from the input file using pandas with tab as the separator
    data = pd.read_csv(input_file, sep='\t')

    # Read combined_annotations
    combined_data = pd.read_csv(combined_annotations, sep='\t')

    # Deduplicate and extract unique sample values
    unique_samples = combined_data['sample'].unique()

    # Create a Workbook
    wb = Workbook()

    # Create the genome_stats sheet
    gs_sheet = wb.create_sheet(title="genome_stats")

    # Dynamically get unique RNA types from combined_rrna
    rrna_data = pd.read_csv(combined_rrna, sep='\t')
    rna_columns = rrna_data['type'].unique()  # Define rna_columns here
    unique_rna_types = rrna_data['type'].unique()

    # Append column names to genome_stats sheet
    column_names = ["sample", "number of scaffolds", "taxonomy", "completeness", "contamination"] + list(rna_columns)
    gs_sheet.append(column_names)

    # Populate genome_stats sheet with data from combined_annotations
    for sample in unique_samples:
        # Extract information for the current sample from combined_annotations
        sample_info = combined_data[combined_data['sample'] == sample].iloc[0]  # Assuming one row per sample

        # Append data to genome_stats sheet
        gs_sheet.append([sample, None, sample_info.get('taxonomy', None), sample_info.get('Completeness', None), sample_info.get('Contamination', None)] + [None] * 3)

    # Initialize a list for tRNA count values
    tRNA_count_values = [0] * len(unique_samples)

    # Initialize lists for completeness and contamination values
    completeness_values = [0] * len(unique_samples)
    contamination_values = [0] * len(unique_samples)

    # Calculate "tRNA count" and other values for each sample
    trna_data = pd.read_csv(trna_file, sep='\t')
    sample_names = trna_data.columns[5:]  # Assuming the relevant columns start from index 5

    for idx, sample in enumerate(unique_samples):
        tRNA_count_values[idx] = trna_data[sample].sum()

        # Extract completeness and contamination values from combined_annotations
        sample_info = combined_data[combined_data['sample'] == sample].iloc[0]  # Assuming one row per sample
        completeness_values[idx] = sample_info.get('Completeness', None)
        contamination_values[idx] = sample_info.get('Contamination', None)

    # Add "tRNA count" column to genome_stats sheet as the last column
    gs_sheet.insert_cols(len(gs_sheet[1]) + 1, amount=1)  # Insert a new column as the last column
    gs_sheet.cell(row=1, column=len(gs_sheet[1]), value="tRNA count")  # Set the column header

    # Populate "tRNA count" column with values for each sample
    for idx, value in enumerate(tRNA_count_values, start=2):
        gs_sheet.cell(row=idx, column=len(gs_sheet[1]), value=value)

    # Update RNA columns dynamically
    for rna_type in unique_rna_types:
        # Filter rrna_data based on rna_type
        filtered_rrna_data = rrna_data[rrna_data['type'] == rna_type]

        # Find the corresponding column index in the genome_stats sheet
        col_idx = column_names.index(rna_type) + 1
        print(f"\nUpdating RNA column: {rna_type}")
        print(f"Column Index in genome_stats: {col_idx}")

        # Iterate over samples
        for idx, sample in enumerate(unique_samples, start=2):
            # Extract relevant data for the current sample and rna_type
            sample_rrna_data = filtered_rrna_data[filtered_rrna_data['sample'] == sample]

            # Check if sample_rrna_data is not empty
            if not sample_rrna_data.empty:
                # Concatenate the values and format them as needed
                values = [
                    f"{row['query_id']}, ({row['begin']}, {row['end']})"
                    for _, row in sample_rrna_data.iterrows()
                ]

                # Join multiple values with "; "
                joined_values = "; ".join(values)

                # Find the corresponding row index for the sample
                for row in gs_sheet.iter_rows(min_row=2, max_row=gs_sheet.max_row, min_col=1, max_col=1):
                    if row[0].value == sample:
                        row_idx = row[0].row
                        break

                # Update the value in the correct cell
                if rna_type == 'tRNA':
                    gs_sheet.cell(row=row_idx, column=len(gs_sheet[1])).value = joined_values
                    print(f"Updated tRNA count at row {row_idx}, column {len(gs_sheet[1])}")
                else:
                    gs_sheet.cell(row=row_idx, column=col_idx).value = joined_values
                    print(f"Updated value at row {row_idx}, column {col_idx}")

    # Print completeness, contamination, and RNA values
    print("\nCompleteness, Contamination, and RNA values:")
    for row in gs_sheet.iter_rows(min_row=2, max_row=gs_sheet.max_row, min_col=4, values_only=True):
        print(row)

    print("\nUpdated Genome Stats Sheet:")
    for row in gs_sheet.iter_rows(min_row=1, max_row=gs_sheet.max_row, values_only=True):
        print(row)

    # Create a dictionary to store data for each sheet
    sheet_data = {}

    # Fixed columns
    fixed_columns = ['gene_id', 'gene_description', 'pathway', 'topic_ecosystem', 'category', 'subcategory']

    for _, row in data.iterrows():
        # Split the "sheet" values by "; " and iterate over them
        for sheet_name in row['topic_ecosystem'].split('; '):  # Assuming 'topic_ecosystem' corresponds to 'sheet'
            sheet_name = sheet_name.replace(" ", "_")
            if sheet_name not in sheet_data:
                sheet_data[sheet_name] = []

            # Exclude the "sheet" column and move "gene_id" as the second column
            row_data = [row[col] for col in fixed_columns]

            # Include the 'potential_amg' column if it exists
            if 'potential_amg' in data.columns:
                # Convert 'potential_amg' values to "TRUE" or "FALSE"
                row_data += ['TRUE' if row['potential_amg'] == 'TRUE' else 'FALSE']

            # Append the rest of the columns without 'potential_amg'
            row_data += [row[col] for col in data.columns if col not in fixed_columns and col != 'potential_amg']

            # Append the modified row to the corresponding sheet
            sheet_data[sheet_name].append(row_data)

    for sheet_name, sheet_rows in sheet_data.items():
        # Create a worksheet for each sheet
        ws = wb.create_sheet(title=sheet_name)

        # Extract column names from the original DataFrame
        column_names = fixed_columns

        # Include 'potential_amg' column if it exists
        if 'potential_amg' in data.columns:
            column_names += ['potential_amg']

        # Append the rest of the columns without 'potential_amg'
        column_names += [col for col in data.columns if col not in fixed_columns and col != 'potential_amg']

        # Append column names as the first row
        ws.append(column_names)

        # Append data rows to the worksheet
        for r_idx, row in enumerate(sheet_rows, 1):
            ws.append(row)

        # Create a table from the data for filtering
        tab = Table(displayName=f"{sheet_name}_Table", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Add rRNA and tRNA sheets
    rrna_data = pd.read_csv(rrna_file, sep='\t')
    trna_data = pd.read_csv(trna_file, sep='\t')

    rrna_sheet = wb.create_sheet(title="rRNA")
    trna_sheet = wb.create_sheet(title="tRNA")

    for sheet, data in zip([rrna_sheet, trna_sheet], [rrna_data, trna_data]):
        # Append column names as the first row
        sheet.append(list(data.columns))

        # Append data rows to the worksheet
        for _, row in data.iterrows():
            sheet.append(list(row))

    # Remove the default "Sheet" that was created
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    # Save the workbook as the output file
    wb.save(output_file)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate multi-sheet XLSX file')
    parser.add_argument('--input-file', required=True, help='Path to the input TSV file')
    parser.add_argument('--rrna-file', required=True, help='Path to the rRNA TSV file')
    parser.add_argument('--trna-file', required=True, help='Path to the tRNA TSV file')
    parser.add_argument('--combined-annotations', required=True, help='Path to the combined_annotations TSV file')
    parser.add_argument('--combined-rrna', required=True, help='Path to the combined_rrna TSV file')
    parser.add_argument('--output-file', required=True, help='Path to the output XLSX file')

    args = parser.parse_args()
    generate_multi_sheet_xlsx(args.input_file, args.rrna_file, args.trna_file, args.combined_annotations, args.combined_rrna, args.output_file)
