import pandas as pd
import sqlite3
import argparse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# Setup logging to display messages to console
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_arguments():
    parser = argparse.ArgumentParser(description='Generate a multi-sheet XLSX document from distill sheets and a SQLite database.')
    parser.add_argument('--target_id_counts', type=str, help='Path to the target_id_counts.tsv file.')
    parser.add_argument('--db_name', type=str, help='Name of the SQLite database file.')
    parser.add_argument('--distill_sheets', nargs='+', help='List of paths to distill sheets.')
    parser.add_argument('--rrna_file', type=str, help='Path to the rrna_sheet.tsv file.', default=None)
    parser.add_argument('--combined_rrna_file', type=str, help='Path to the combined rRNA TSV file.', default=None)
    parser.add_argument('--trna_file', type=str, help='Path to the trna_sheet.tsv file.', default=None)
    parser.add_argument('--output_file', type=str, help='Path to the output XLSX file.')
    return parser.parse_args()

def compile_target_id_counts(target_id_counts):
    return pd.read_csv(target_id_counts, sep='\t')

import logging

def read_distill_sheets(distill_sheets):
    sheets_data = {}
    for sheet_path in distill_sheets:
        if file_contains_data(sheet_path):
            df = pd.read_csv(sheet_path, sep='\t')
            topic = df['topic_ecosystem'].unique().tolist()
            column_type = 'ec_id' if 'ec_id' in df.columns else 'gene_id'
            logging.debug(f"For sheet {sheet_path}, identified column type as: {column_type}")
            sheets_data.update({sheet_path: {'dataframe': df, 'topics': topic, 'column_type': column_type}})
        else:
            print(f"Skipping {sheet_path} as it contains 'NULL'.")
    return sheets_data


def file_contains_data(file_path):
    try:
        with open(file_path, 'r') as f:
            first_line = f.readline().strip()
            return first_line != "NULL"
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return False

def compile_genome_stats(db_name):
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    
    # Check for the presence of optional columns in the database
    cursor.execute("PRAGMA table_info(annotations)")
    columns_info = cursor.fetchall()
    column_names = [info[1] for info in columns_info]
    
    # Start with a query to select distinct samples
    query_base = "SELECT DISTINCT sample FROM annotations"
    df_samples = pd.read_sql_query(query_base, conn)
    
    # Initialize the genome_stats DataFrame with sample names
    df_genome_stats = pd.DataFrame({"sample": df_samples['sample']})
    
    # Dynamically build a query to include optional columns if they are present
    optional_columns = ['taxonomy', 'Completeness', 'Contamination']
    select_clauses = []
    for col in optional_columns:
        if col in column_names:
            # Prepare a SELECT clause to calculate average for numeric columns and group_concat for taxonomy
            if col in ['Completeness', 'Contamination']:
                select_clauses.append(f"AVG({col}) AS {col}")
            else:  # Assuming 'taxonomy' is not a numeric column
                select_clauses.append(f"GROUP_CONCAT(DISTINCT {col}) AS {col}")
    
    # If there are optional columns to include, modify the base query
    if select_clauses:
        query = f"SELECT sample, {', '.join(select_clauses)} FROM annotations GROUP BY sample"
        df_stats = pd.read_sql_query(query, conn)
        df_genome_stats = pd.merge(df_genome_stats, df_stats, on="sample", how="left")
    else:
        # If no optional columns, the genome_stats will only contain sample names at this point
        pass
    
    conn.close()
    return df_genome_stats

def add_sheet_from_dataframe(wb, df, sheet_name):
    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


def prepare_ec_like_patterns(ec_number):
    """
    Prepare SQL LIKE patterns for partial matching of EC numbers.
    Handles cases where EC numbers are partial or contain multiple EC numbers separated by semicolons or spaces.
    """
    patterns = []
    # Split multiple EC numbers and prepare patterns for each
    parts = ec_number.replace(" ", ";").split(";")  # Split by semicolon and space
    for part in parts:
        clean_part = part.strip().rstrip('.')
        if clean_part:  # Ensure the part is not empty after stripping
            # Append '%' to match any characters following the specified EC number part
            pattern = clean_part + "%"
            patterns.append(pattern)
    logging.debug(f"Generated patterns for EC number {ec_number}: {patterns}")
    return patterns

# Used for partial EC matching - good in theory but can easily produce too many hits for XLSX
# def query_annotations_for_gene_ids(db_name, ids, column_type):
#     conn = sqlite3.connect(db_name)
#     df_result = pd.DataFrame()

#     for id_value in ids:
#         logging.debug(f"Processing ID: {id_value} as {column_type}")
#         if column_type == 'ec_id':
#             like_patterns = prepare_ec_like_patterns(id_value)
#             for pattern in like_patterns:
#                 query = "SELECT DISTINCT gene_id FROM annotations WHERE gene_id LIKE ? ESCAPE '\\'"
#                 logging.debug(f"Executing query with pattern: {pattern}")
#                 df_partial = pd.read_sql_query(query, conn, params=(pattern,))
#                 if not df_partial.empty:
#                     logging.debug(f"Found matches for pattern {pattern}: {df_partial['gene_id'].tolist()}")
#                 df_result = pd.concat([df_result, df_partial], ignore_index=True)
#         else:
#             query = "SELECT DISTINCT gene_id FROM annotations WHERE gene_id = ?"
#             logging.debug(f"Executing query for exact match: {id_value}")
#             df_partial = pd.read_sql_query(query, conn, params=(id_value,))
#             df_result = pd.concat([df_result, df_partial], ignore_index=True)

#     df_result.drop_duplicates(inplace=True)
#     conn.close()
#     logging.debug(f"Final matched gene_ids: {df_result['gene_id'].tolist()}")
#     return df_result

def query_annotations_for_gene_ids(db_name, ids, column_type):
    conn = sqlite3.connect(db_name)
    df_result = pd.DataFrame()

    # Fetch all gene_id from annotations to filter in Python (for demonstration, may need optimization)
    all_gene_ids = pd.read_sql_query("SELECT DISTINCT gene_id FROM annotations", conn)

    for id_value in ids:
        logging.debug(f"Processing ID: {id_value} as {column_type}")
        # Directly match for gene_id
        if column_type == 'gene_id':
            df_partial = all_gene_ids[all_gene_ids['gene_id'] == id_value]
        elif column_type == 'ec_id':
            # Filter function to match any EC number within entries
            def match_ec(ec_entry):
                # Split by both semicolon and space, then check if id_value matches any
                split_ecs = re.split('; | ', ec_entry)  # Adjust regex as needed
                return any(id_value == ec for ec in split_ecs)

            df_partial = all_gene_ids[all_gene_ids['gene_id'].apply(match_ec)]

        if not df_partial.empty:
            logging.debug(f"Found matches for ID {id_value}: {df_partial['gene_id'].tolist()}")
        df_result = pd.concat([df_result, df_partial], ignore_index=True)

    df_result.drop_duplicates(inplace=True)
    conn.close()
    logging.debug(f"Final matched gene_ids: {df_result['gene_id'].tolist()}")
    return df_result

def compile_rrna_information(combined_rrna_file):
    """Compile rRNA information from the combined rRNA file."""
    rrna_data = pd.read_csv(combined_rrna_file, sep='\t')
    # Group by sample and type to concatenate query_id and positions
    rrna_summary = rrna_data.groupby(['sample', 'type']).apply(
        lambda x: '; '.join([f"{row['query_id']} ({row['begin']}, {row['end']})" for _, row in x.iterrows()])
    ).unstack(fill_value='')
    rrna_summary.reset_index(inplace=True)
    return rrna_summary

def add_rrna_trna_sheets(wb, rrna_file, trna_file):
    for tsv_file, sheet_name in [(rrna_file, 'rRNA'), (trna_file, 'tRNA')]:
        if tsv_file and file_contains_data(tsv_file):
            df = pd.read_csv(tsv_file, sep='\t')
            add_sheet_from_dataframe(wb, df, sheet_name)

def compile_rrna_data(rrna_file):
    """Compile rRNA data from the given file."""
    rrna_data = pd.read_csv(rrna_file, sep='\t')
    # Process rRNA data as required for your use case
    # This might involve summarizing the rRNA types and locations for each sample
    return rrna_data

def compile_trna_counts(trna_file):
    trna_data = pd.read_csv(trna_file, sep='\t')
    sample_columns = trna_data.columns[5:]  # Adjust based on actual structure
    trna_counts = pd.DataFrame([{'sample': sample, 'tRNA count': trna_data[sample].sum()} for sample in sample_columns])
    return trna_counts

def update_genome_stats_with_rrna_trna(genome_stats_df, rrna_file, trna_file):
    """Update the genome_stats DataFrame with rRNA and tRNA information."""
    # Process rRNA file if it contains data
    if file_contains_data(rrna_file):
        rrna_data = compile_rrna_data(rrna_file)
        # Integrate rrna_data into genome_stats_df as needed

    # Process tRNA file if it contains data
    if file_contains_data(trna_file):
        trna_counts = compile_trna_counts(trna_file)
        # Merge tRNA counts into genome_stats_df
        genome_stats_df = pd.merge(genome_stats_df, trna_counts, on="sample", how="left")

    return genome_stats_df

def update_genome_stats_with_rrna(genome_stats_df, combined_rrna_file):
    """Update the genome_stats DataFrame with rRNA information if available."""
    if file_contains_data(combined_rrna_file):
        rrna_summary = compile_rrna_information(combined_rrna_file)
        genome_stats_df = pd.merge(genome_stats_df, rrna_summary, on="sample", how="left")
    return genome_stats_df

def main():
    args = parse_arguments()

    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Compile genome stats and add as a sheet
    genome_stats_df = compile_genome_stats(args.db_name)
    genome_stats_df = update_genome_stats_with_rrna_trna(genome_stats_df, args.rrna_file, args.trna_file)
    genome_stats_df = update_genome_stats_with_rrna(genome_stats_df, args.combined_rrna_file)
    add_sheet_from_dataframe(wb, genome_stats_df, "Genome_Stats")

    # Read target ID counts
    target_id_counts_df = compile_target_id_counts(args.target_id_counts)
    
    # Process each distill sheet
    distill_data = read_distill_sheets(args.distill_sheets)
    for sheet_path, info in distill_data.items():
        df_distill = info['dataframe']
        column_type = info['column_type']
        
        # Iterate through topics within each distill sheet
        for topic in info['topics']:
            df_topic = df_distill[df_distill['topic_ecosystem'] == topic]
            gene_ids = df_topic[column_type].unique().tolist()
            
            # Query annotations database for each gene_id/ec_id to find matches
            df_matched_gene_ids = query_annotations_for_gene_ids(args.db_name, gene_ids, column_type)
            
            # Merge matched gene IDs with target ID counts
            df_merged = pd.merge(df_topic, df_matched_gene_ids, left_on=column_type, right_on='gene_id', how='inner')
            df_merged_with_counts = pd.merge(df_merged, target_id_counts_df, on='gene_id', how='left')
            
            # Drop unnecessary columns
            df_final = df_merged_with_counts.drop(columns=['query_id', 'sample', 'taxonomy', 'Completeness', 'Contamination'], errors='ignore')
            
            # Rename column_type to gene_id if it's ec_id for consistency in output
            if column_type == 'ec_id':
                df_final.rename(columns={'ec_id': 'gene_id'}, inplace=True)
            
            # Add final dataframe as a new sheet in the workbook
            sheet_name = topic[:31]  # Limit sheet name to 31 characters
            add_sheet_from_dataframe(wb, df_final, sheet_name)

    # Add rRNA and tRNA sheets if available
    if args.rrna_file and file_contains_data(args.rrna_file):
        df_rrna = pd.read_csv(args.rrna_file, sep='\t')
        add_sheet_from_dataframe(wb, df_rrna, 'rRNA')
    
    if args.trna_file and file_contains_data(args.trna_file):
        df_trna = pd.read_csv(args.trna_file, sep='\t')
        add_sheet_from_dataframe(wb, df_trna, 'tRNA')

    # Save the workbook
    wb.save(args.output_file)

# Used for partial EC matching - good in theory but can easily produce too many hits for XLSX
# def expand_df_with_matched_gene_ids(df, gene_ids, column_type, db_name):
#     expanded_rows = []
#     for gene_id in gene_ids:
#         matched_gene_ids = query_annotations_for_gene_ids(db_name, [gene_id], column_type)
#         for matched_gene_id in matched_gene_ids['gene_id'].tolist():
#             for _, row in df.iterrows():
#                 new_row = row.copy()
#                 new_row['gene_id'] = matched_gene_id  # Update gene_id with matched value
#                 expanded_rows.append(new_row)
#     return pd.DataFrame(expanded_rows)

# Used for partial EC matching - good in theory but can easily produce too many hits for XLSX
# def main():
#     args = parse_arguments()

#     wb = Workbook()
#     wb.remove(wb.active)  # Remove the default sheet

#     genome_stats_df = compile_genome_stats(args.db_name)
#     # Update genome_stats with rRNA and tRNA data if available
#     genome_stats_df = update_genome_stats_with_rrna_trna(genome_stats_df, args.rrna_file, args.trna_file)
#     genome_stats_df = update_genome_stats_with_rrna(genome_stats_df, args.combined_rrna_file)
#     add_sheet_from_dataframe(wb, genome_stats_df, "genome_stats")

#     target_id_counts_df = compile_target_id_counts(args.target_id_counts)
#     distill_data = read_distill_sheets(args.distill_sheets)
#     for sheet_path, info in distill_data.items():
#         df_distill = info['dataframe']
#         column_type = info['column_type']  # Extract column_type here
#         for topic in info['topics']:
#             df_topic = df_distill[df_distill['topic_ecosystem'] == topic]
#             gene_ids = df_topic[column_type].unique().tolist()  # Use column_type to extract IDs

#             # Adjusted to pass column_type
#             df_valid_gene_ids = query_annotations_for_gene_ids(args.db_name, gene_ids, column_type)

#             # Filter the df_topic to keep only rows where gene_id exists in annotations.db
#             # Instead of accessing 'gene_id' directly, use the column_type attribute
#             df_topic_filtered = df_topic[df_topic[info['column_type']].isin(df_valid_gene_ids['gene_id'])].copy()

#             # Instead of accessing 'gene_id' directly, use the column_type attribute
#             column_type = distill_data[sheet_path]['column_type']
#             # Instead of directly renaming and merging df_topic_filtered
#             if column_type == 'ec_id':
#                 logging.debug(f"Columns before renaming: {df_topic_filtered.columns}")
#                 # No need to rename here as it will be handled in the expand_df_with_matched_gene_ids function
#                 logging.debug(f"Columns after renaming: {df_topic_filtered.columns}")

#             # Fetch the matched gene_ids for the entire set of EC numbers or gene_ids
#             matched_gene_ids = query_annotations_for_gene_ids(args.db_name, gene_ids, column_type)

#             # Expand df_topic_filtered to include a row for each matched gene_id
#             df_topic_filtered_expanded = expand_df_with_matched_gene_ids(df_topic_filtered, gene_ids, column_type, args.db_name)

#             logging.debug(f"df_topic_filtered_expanded gene_id unique values: {df_topic_filtered_expanded['gene_id'].unique()}")
#             logging.debug(f"df_topic_filtered_expanded shape: {df_topic_filtered_expanded.shape}")

#             # Proceed with the merge operation using the expanded dataframe
#             df_merged = pd.merge(df_topic_filtered_expanded, target_id_counts_df, on='gene_id', how="left")

#             logging.debug(f"df_merged shape after merge: {df_merged.shape}")

#             df_final = df_merged.drop(columns=['query_id', 'sample', 'taxonomy', 'Completeness', 'Contamination'], errors='ignore')
            
#             sheet_name = topic[:31]  # Excel sheet name character limit
#             add_sheet_from_dataframe(wb, df_final, sheet_name)

#     # Add rrna and trna sheets if not NULL
#     add_rrna_trna_sheets(wb, args.rrna_file, args.trna_file)

#     wb.save(args.output_file)

if __name__ == '__main__':
    main()
